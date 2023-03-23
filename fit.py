from flask import *
from openpyxl import *

XLSX_FILENAME = "fit.xlsx"

wb = load_workbook(filename = XLSX_FILENAME)
inventory = wb["Inventory"]
rooms = wb["Rooms"]
assets = wb["Assets"]
log = wb["Log"]

def collect_sheet(sheet):
	header = next(sheet.rows)
	body = sheet.iter_rows(min_row=2, max_col=len(header))
	return {int(row[0].value): {key.value: cell.value for (key, cell) in zip(header, row)} for row in body}

def sort_sheet(sheet):
	D = collect_sheet(sheet)
	return [D[k] for k in sorted(D.keys())]

def sheet_hash(sorted_sheet, pk):
	return {sorted_sheet[i][pk]: i for i in range(len(sorted_sheet))}

def next_available_id(sheet, pk):
	try:
		return max([int(id) for id in [entry[pk] for entry in sort_sheet(sheet)]]) + 1
	except ValueError:
		return 1

def row_index(sheet, pk, target):
	s = collect_sheet(sheet)
	for row, i in zip(s.values(), range(len(s))):
		if row[pk] == target:
			return i + 1
	raise KeyError(f"primary key \"{target}\" not found")

def _move_asset(asset_id, src_room_id, dest_room_id, quantity):
	# Check if IDs are valid. Raise KeyError if not.
	row_index(assets, "Asset ID", asset_id)
	row_index(rooms, "Room ID", src_room_id)
	row_index(rooms, "Room ID", dest_room_id)
	
	header = [cell.value for cell in next(inventory.rows)]
	ROOM_ID = header.index("Room ID")
	ASSET_ID = header.index("Asset ID")
	QUANTITY = header.index("Quantity")

	for i in range(2, len(inventory.rows) + 1):
		rid = inventory.cell(row = i, column = ROOM_ID).value
		aid = inventory.cell(row = i, column = ASSET_ID).value
		if aid == asset_id:
			if rid == src_room_id:
				qu_cell = inventory.cell(row = i, column = QUANTITY)
				qu_cell.value = int(qu_cell.value) - quantity
			if rid == dest_room_id:
				qu_cell = inventory.cell(row = i, column = QUANTITY)
				qu_cell.value = int(qu_cell.value) + quantity

def _log(operation, details):
	log.append([ next_available_id(log, "Log ID"), operation, details ])

def _save_wb():
	wb.save(filename = XLSX_FILENAME)


app = Flask(__name__)

@app.route("/")
def flask_root():
	return ""  # TODO: stub

@app.route("/asset", methods=["GET"])
def get_assets():
	return collect_sheet(assets)

@app.route("/asset", methods=["POST"])
def create_asset():
	POST = request.form
	if ("asset_name" not in POST):
		abort(400)
	
	asset_id = next_available_id(assets, "Asset ID")
	asset_name = POST.get("asset_name")
	product_number = POST.get("product_number")
	manufacturer = POST.get("manufacturer")
	assets.append([asset_id, asset_name, product_number, manufacturer])
	_log("create_asset", \
		f"asset_id={asset_id}; asset_name={asset_name}; product_number={product_number}; manufacturer={manufacturer}")
	_save_wb()
	return asset_id

@app.route("/asset/<int:id>", methods=["GET"])
def asset_info(id):
	try:
		return collect_sheet(assets)[int(id)]
	except KeyError:
		abort(404)

@app.route("/asset/<int:id>", methods=["PUT"])
def update_asset_info(id):
	PUT = request.form	
	try:
		row_idx = row_index(assets, "Asset ID", str(id)) + 1  # may raise KeyError
		assets.cell(row = row_idx, column = 2).value = PUT["asset_name"]  # may raise KeyError
		assets.cell(row = row_idx, column = 3).value = PUT.get("product_number")
		assets.cell(row = row_idx, column = 4).value = PUT.get("manufacturer")
		_log("update_asset_info", \
			f"id={id}; asset_name={PUT['asset_name']}; product_number={PUT.get('product_number')}; manufacturer={PUT.get('manufacturer')}")
		_save_wb()
	except KeyError:
		abort(404)

@app.route("/asset/<int:id>", methods=["DELETE"])
def remove_asset(id):
	DELETE = request.args
	# if "erase" in DELETE.keys():  # full erasure of asset vs. move all quantities to __trash
	abort(500)  # TODO stub

@app.route("/room", methods=["GET"])
def get_rooms():
	return collect_sheet(rooms)

@app.route("/room", methods=["POST"])
def create_room():
	POST = request.form
	if ("room_name" not in POST):
		abort(400)
	
	room_id = next_available_id(rooms, "Room ID")
	room_name = POST.get("room_name")
	description = POST.get("description")
	staff = POST.get("staff")
	assets.append([room_id, room_name, description, staff])
	_log("create_room", \
		f"room_id={room_id}; room_name={room_name}; description={description}; staff={staff}")
	_save_wb()

@app.route("/room/<int:id>", methods=["GET"])
def room_info(id):
	try:
		return collect_sheet(rooms)[int(id)]
	except KeyError:
		abort(404)

@app.route("/room/<int:id>", methods=["PUT"])
def update_room_info(id):
	PUT = request.form	
	try:
		row_idx = row_index(rooms, "Room ID", str(id)) + 1  # may raise KeyError
		rooms.cell(row = row_idx, column = 2).value = PUT["room_name"]  # may raise KeyError
		rooms.cell(row = row_idx, column = 3).value = PUT.get("description")
		rooms.cell(row = row_idx, column = 4).value = PUT.get("staff")
		# TODO: log vvvvv
		_save_wb()
	except KeyError:
		abort(404)

@app.route("/room/<int:id>", methods=["DELETE"])
def delete_room():
	pass  # TODO: stub

@app.route("/move", methods=["POST"])
def move_asset():
	POST = request.form
	try:
		asset = POST.get("asset")
		src = POST.get("src")
		dest = POST.get("dest")
		qu = POST.get("qu")

		_move_asset(asset, src, dest, qu)
		_log("move", f"asset={asset}; src={src}; dest={dest}; qu={qu}")
		_save_wb()
	
	except KeyError:
		abort(400)

@app.route("/split", methods=["POST"])
def split_asset():
	pass  # TODO: stub

@app.route("/merge", methods=["POST"])
def merge_assets():
	pass  # TODO: stub

@app.route("/log", methods=["GET"])
def get_log_list():
	return collect_sheet(log)

@app.route("/log/<int:id>", methods=["GET"])
def get_log_entry(id):
	return collect_sheet(log)[id]

@app.route("/log/<int:id>", methods=["DELETE"])
def undo_log_entry(id):
	pass  # TODO: stub

@app.route("/inventory", methods=["GET"])
def inventory_table():
	as_list = sort_sheet(assets)
	rm_list = sort_sheet(rooms)
	as_hash = sheet_hash(as_list, "Asset ID")
	rm_hash = sheet_hash(rm_list, "Room ID")

	table = [["Room"] + [a["Asset Name"] for a in as_list]]
	table.extend([[[r["Room Name"]] + [0] * len(as_list)] for r in rm_list])
	
	header = [cell.value for cell in next(inventory.rows)]
	ROOM_ID = header.index("Room ID")
	ASSET_ID = header.index("Asset ID")
	QUANTITY = header.index("Quantity")

	for row in inventory.iter_rows(min_row=2):
		rm_idx = rm_hash[row[ROOM_ID]]
		as_idx = as_hash[row[ASSET_ID]]
		table[rm_idx][as_idx] += int(row[QUANTITY])
	
	return table
